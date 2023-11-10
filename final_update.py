import pandas as pd
import numpy as np
import openpyxl as op
from openpyxl.styles import colors, Border, Alignment, Side
from subjects_EC import subjectsDict
import os
from copy import copy
import re

# 定义字体样式
fontOld = op.styles.Font(name='Times New Roman', size=11, color='000000')  # 黑色
fontOldBold = op.styles.Font(name='Times New Roman', size=11, bold=True, color='000000')  # 黑色加粗
fontNew = op.styles.Font(name='Times New Roman', size=11, color='FF0000')  # 红色
fontNewBold = op.styles.Font(name='Times New Roman', size=11, bold=True, color='FF0000')  # 红色加粗
fontBold = op.styles.Font(name='Times New Roman', bold=True)  # 黑色加粗

# 定义边框样式
borderSet = Border(left=Side(style='thin', color=colors.BLACK),
                   right=Side(style='thin', color=colors.BLACK),
                   top=Side(style='thin', color=colors.BLACK),
                   bottom=Side(style='thin', color=colors.BLACK))
# 定义对齐格式
alignmentSet = Alignment(horizontal='center', vertical='center', wrap_text=True)


def update_style(ws, row, new_col):
    """将左侧旧数据单元格格式设置为黑色，并将右侧新数据格式设置为红色"""
    newCell = ws.cell(row=row, column=new_col)
    oldCell = ws.cell(row=row, column=new_col - 1)
    exCell = ws.cell(row=row, column=new_col - 2)
    newCell._style = copy(oldCell._style)
    oldCell._style = copy(exCell._style)


class StatisticUpdate:
    def __init__(self, datas_file, target_file, res_path, year, month):
        self.df_data = pd.read_excel(datas_file, sheet_name='中国高校')
        self.ws_target = op.load_workbook(target_file)
        self.targetFile = target_file
        self.resPath = res_path
        self.levelList = ['百分之一', '千分之一', '万分之一']
        self.year = year
        self.month = month
        self.dateM = int(str(year) + str(month).zfill(2))
        self.collegeDict = None
        self.resultXlsFile = None

    def start(self):
        """运行更新ESI学科统计表程序"""
        try:
            yield "正在更新各省份统计数据..."
            self.province_update()
            yield "各省份统计数据更新完成"
            yield "正在更新上海高校统计数据..."
            self.shanghai_update()
            yield "上海高校统计数据更新完成"
            yield "正在更新北京，上海，江苏三省份学科详细数据表"
            self.update_prov()
            yield "学科详细数据表更新完成"
            self.save_xls()
            yield "finished_upd"
        except Exception:
            yield "update_error"

    def province_update(self):
        """更新‘分省份’sheet"""
        provTargetCol = (self.year - 2017) * 6 + 3 + int((self.month + 1) / 2)  # 在结果统计表中，每一种分档所占的列数
        # 获取省份列表
        df_province = pd.read_excel(self.targetFile, sheet_name='分省份', index_col=0, skiprows=1, skipfooter=3)
        provinceList = list(df_province.index)[:-1]
        self.datas_update('分省份', provTargetCol, provinceList)
    
    def shanghai_update(self):
        """更新‘上海高校’sheet"""
        SHTargetCol = (self.year - 2017) * 6 + 1 + int((self.month + 1) / 2)  # 在结果统计表中，每一种分档所占的列数
        # 获取上海高校列表
        df_collegeSH = pd.read_excel(self.targetFile, sheet_name='上海高校', index_col=0, skiprows=1, skipfooter=3)
        collegeSHList = list(df_collegeSH.index)[:-1]
        self.datas_update('上海高校', SHTargetCol, collegeSHList)

    def datas_update(self, sheet_name, target_col, indexList):
        """工作表数据更新"""
        # 百分之一、千分之一、万分之一的最后一列所对应的列编号
        hundEnd = target_col
        thouEnd = target_col * 2 - 1
        tenthouEnd = target_col * 3 - 2
        endColList = [1, hundEnd, thouEnd, tenthouEnd]

        # 将各个省份的三种分档学科数
        df_hund = self.df_data[self.df_data['FIELD'] != 'ALL']
        if sheet_name == '上海高校':
            # 若更新的表为上海高校，则需要再提取出上海市数据
            df_hund = df_hund[df_hund['省份'] == '上海']

        df_thou = df_hund[df_hund['分档'] != '百分之一']
        df_tenthou = df_hund[df_hund['分档'] == '万分之一']
        dfList = [df_hund, df_thou, df_tenthou]

        # 将三个分档的学科数据存于字典sr_dataProvince中
        sr_data = dict()
        for i, level in enumerate(self.levelList):
            targetHead = '高校' if sheet_name == '上海高校' else '省份'
            sr_data[level] = pd.pivot_table(dfList[i], index=[targetHead], aggfunc='size')

        ws_target = self.ws_target[sheet_name]  # 通过openpyxl读取结果统计表

        # 判断是否有新增高校，若有则在结果行添加该项目数据条
        oldIndexList = indexList
        newIndexList = list(sr_data['百分之一'].index)
        if len(newIndexList) > len(oldIndexList):
            addList = list(set(newIndexList).difference(set(oldIndexList)))
            indexList = indexList + list(reversed(addList))  # 将新增项目添加至indexList
            addRow = len(oldIndexList) + 3
            for item in addList:
                ws_target.insert_rows(idx=addRow)
                ws_target.cell(row=addRow, column=1).value = item
                ws_target.cell(row=addRow, column=1).alignment = alignmentSet
                # 修改当前添加行的格式
                for c in range(tenthouEnd):
                    ws_target.cell(row=addRow, column=c + 1)._style = copy(
                        ws_target.cell(row=addRow - 1, column=c + 1)._style)

        # 将结果统计表第一行单元格取消合并
        for i in range(len(endColList) - 1):
            ws_target.unmerge_cells(start_row=1,
                                    start_column=endColList[i] + 1,
                                    end_row=1,
                                    end_column=endColList[i + 1]
                                    )

        # 进行结果统计表数据更新操作
        for i in range(len(endColList) - 1):

            newCol = endColList[i + 1] + i + 1  # 插入更新列后的该分档下最后一列列编号
            ws_target.insert_cols(idx=newCol)  # 插入更新列

            # 插入更新列的表头并修改格式
            ws_target.cell(row=2, column=newCol).value = self.dateM
            update_style(ws_target, 2, newCol)

            levelDatas = sr_data[self.levelList[i]]  # 获取该分档对应的各省份学科数据表

            # 补充各省份数据并更改单元格格式
            for j, item in enumerate(indexList):
                ws_target.cell(row=j + 3, column=newCol).value = levelDatas.get(item)
                update_style(ws_target, j + 3, newCol)

            # 补充最后一行“总计”的数据并更改单元格格式
            totalRow = len(indexList) + 3
            ws_target.cell(row=totalRow, column=newCol).value = levelDatas.sum()
            update_style(ws_target, totalRow, newCol)

            # 结果统计表第一行单元格合并
            ws_target.merge_cells(start_row=1,
                                  start_column=endColList[i] + i + 1,
                                  end_row=1,
                                  end_column=endColList[i + 1] + i + 1
                                  )

    def update_subject_detail(self, prov):
        """更新省份ESI学科详细统计表"""
        def set_head_style(cell):
            cell.border = borderSet
            cell.alignment = alignmentSet
            cell.font = fontBold

        # table_target = op.load_workbook('./update_test/201609-202109 ESI统计学科数（发市教委）.xlsx')

        ws_prov = self.ws_target[prov]
        prevRow = ws_prov.max_row  # 原统计表的最后一行行号

        # 获得oldList，即上一次更新时的高校名单
        row = prevRow
        oldList = []
        while True:
            # 由于写有“高校”的单元格为合并单元格，该单元格的第二排会被判定为None
            cellValue = ws_prov.cell(row=row, column=1).value
            if cellValue is None:
                break
            oldList.append(cellValue)
            row -= 1

        # 清除上一次更新时的新增学科说明，并获取上一次更新的学科列表
        oldSubsDict = self.get_subs_addition(prov, row + 1, prevRow)

        # 通过ESI数据表提取出更新后的高校名单newList
        df_sub = self.df_data[self.df_data['FIELD'] != 'ALL']
        df_prov = df_sub[df_sub['省份'] == prov]
        newList = list(df_prov['高校'].drop_duplicates(keep='first').values)

        # 根据上一次更新的和本次更新的高校名单获得最新名单indexList
        addList = list(set(newList).difference(set(oldList)))
        indexList = list(reversed(oldList)) + addList  # 将新增项目添加至indexList

        # 填写“高校”表头
        cell = ws_prov.cell(row=prevRow + 2, column=1)
        cell.value = '高校'
        set_head_style(cell)
        ws_prov.merge_cells(start_row=prevRow + 2,
                            start_column=1,
                            end_row=prevRow + 3,
                            end_column=1
                            )

        # 修改中间表头格式（加边框），并输入“合计”
        ws_prov.merge_cells(start_row=prevRow + 2,
                            start_column=2,
                            end_row=prevRow + 3,
                            end_column=2)
        cell = ws_prov.cell(row=prevRow + 2, column=2)
        cell.value = '合计'
        set_head_style(cell)

        # 添加更新时间表头
        cell = ws_prov.cell(row=prevRow + 2, column=3)
        cell.value = self.dateM
        set_head_style(cell)
        ws_prov.merge_cells(start_row=prevRow + 2,
                            start_column=3,
                            end_row=prevRow + 2,
                            end_column=5
                            )
        # 添加分档表头
        levList = ['ESI学科数', '千分之一', '万分之一']
        for i, head in enumerate(levList):
            cell = ws_prov.cell(row=prevRow + 3, column=3 + i)
            cell.value = head
            set_head_style(cell)

        # 填充本次更新的高校于表格中
        for i, item in enumerate(indexList):
            cell = ws_prov.cell(row=prevRow + 4 + i, column=1)
            cell.value = item
            cell.border = borderSet
            cell.alignment = alignmentSet

        # 填写所有学科总数并合并单元格
        cell = ws_prov.cell(row=prevRow + 4, column=2)
        cell.value = df_prov.shape[0]
        cell.border = borderSet
        cell.alignment = alignmentSet
        cell.font = fontNew
        ws_prov.merge_cells(start_row=prevRow + 4,
                            start_column=2,
                            end_row=prevRow + 3 + len(indexList),
                            end_column=2)

        # 填写各所高校对应的学科数
        for i, college in enumerate(indexList):
            # 提取百分之一、千分之一、万分之一的数据表
            dfc_hund = df_prov[df_prov['高校'] == college]
            dfc_thou = dfc_hund[dfc_hund['分档'] != '百分之一']
            dfc_tenthou = dfc_hund[dfc_hund['分档'] == '万分之一']
            dfcList = [dfc_hund, dfc_thou, dfc_tenthou]

            # 填充各个分档的高校学科数据
            for j, dfc in enumerate(dfcList):
                # 修改单元格格式
                cell = ws_prov.cell(row=prevRow + 4 + i, column=j + 3)
                cell.border = borderSet
                cell.alignment = alignmentSet

                # 获取当前高校的所有学科（英文名）列表
                subList_Eng = list(dfc['FIELD'].values)

                # 若该分档下高校无学科，跳过该循环
                if len(subList_Eng) == 0:
                    continue

                # 填充学科中文列表
                subList_Cha = []
                for sub in subList_Eng:
                    subList_Cha.append(subjectsDict[sub])

                oldSubStr = ws_prov.cell(row=prevRow + 4 + i, column=j + 3)
                # subAdd = list(set(subList_Cha).difference(set(oldList)))

                subStr = "、".join(subList_Cha) + '（{:d}）'.format(len(subList_Eng))

                # 对有新增学科的高校添加新增说明
                if college not in oldSubsDict.keys():
                    oldSubList = []
                else:
                    oldSubList = oldSubsDict[college][j]
                addSubs = list(set(subList_Cha).difference(set(oldSubList)))
                redSubs = list(set(oldSubList).difference(set(subList_Cha)))

                if len(addSubs) > 0:
                    subStr += '\n（新增学科：{:s}）'.format('、'.join(addSubs))
                    cell.font = fontNew
                if len(redSubs) > 0:
                    subStr += '\n（减少学科：{:s}）'.format('、'.join(redSubs))
                    cell.font = fontNew

                cell.value = subStr

    def get_subs_addition(self, prov, row1, row2):
        """清除上一次更新时的新增说明，并将所有字体置为黑色"""
        ws = self.ws_target[prov]
        ws.cell(row=row1, column=2).font = fontOld
        subDict = {}
        for i in range(row1, row2 + 1):
            collegeName = ws.cell(row=i, column=1).value
            subDict[collegeName] = []
            for j in range(3, 6):
                subsCell = ws.cell(row=i, column=j)
                if subsCell.value is not None:
                    subsText = re.sub(r'\n（新增学科：.*?）', '', subsCell.value)  # 将单元格的新增学科说明部分去掉
                    subsText = re.sub(r'\n（减少学科：.*?）', '', subsCell.value)  # 将单元格的减少学科说明部分去掉
                else:
                    subsText = ''
                subsCell.value = subsText
                subsCell.font = fontOld

                subListText = re.sub(r'（.*?）', '', subsText)
                subsList = list(subListText.strip().split('、'))
                if len(subsList) == 1 and subsList[0] == '':
                    subsList = []
                subDict[collegeName].append(subsList)
        return subDict

    def update_prov(self):
        """更新北京，江苏，上海三个省份的详细学科表"""
        provList = ['上海', '北京', '江苏']
        for prov in provList:
            self.update_subject_detail(prov)

    def save_xls(self):
        """保存修改好的excel文件"""
        fileName = '201609-{:d} ESI统计学科数（发市教委）.xlsx'.format(self.dateM)
        self.resultXlsFile = os.path.join(self.resPath, fileName)
        self.ws_target.save(self.resultXlsFile)

    def get_excel_filename(self):
        """返回合并后的excel文件路径"""
        return self.resultXlsFile


if __name__ == "__main__":
    datasFile = r'C:\Users\whyyh\Desktop\test\ESI-202205.xlsx'
    targetFile = r'C:\Users\whyyh\Desktop\test\201609-202203 ESI统计学科数（发市教委）.xlsx'
    update = StatisticUpdate(datasFile, targetFile, './', 2022, 5)
    update.province_update()
    update.shanghai_update()
    update.update_prov()
    update.save_xls()
