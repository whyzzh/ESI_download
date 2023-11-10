import time
import pandas as pd
import re
from interface_port import DOCTYPE_NUMBER, FIELD_NUMBER
import openpyxl as op
from openpyxl.styles import Border, Side, colors, Alignment


class DataCombination:
    def __init__(self, excels_path, result_path, year, month):
        """
        初始化结果保存路径
        """
        self.excelsPath = excels_path
        self.resultPath = result_path
        self.contrast_table = pd.read_excel('schools_contrast.xlsx')  # 高校中英文对照表
        self.resultXlsFile = self.resultPath + '\\ESI-{:d}{:0>2d}.xlsx'.format(year, month)

    def start(self):
        """
        数据合并
        """
        comb_all_list = []
        yield "正在合并数据..."
        for i in range(FIELD_NUMBER + 1):
            comb_table = []

            for j in range(DOCTYPE_NUMBER):

                filename = self.excelsPath + '\\%d-%d.xlsx' % (j + 1, i + 1)

                if i == 0:
                    cata = 'ALL'
                else:
                    # 获取当前excel表格的学科类型（“ALL”对应“None”）
                    xls_cata = pd.read_excel(filename, nrows=4)
                    str_cata = xls_cata.iloc[3, 0]
                    cata = re.search('Value\(s\):[\s]*(.*?)[\s]*Show', str_cata).group(1)
                    cata = cata.upper()

                xls_sheet = pd.read_excel(filename, skiprows=5, skipfooter=1)

                if j == 0:
                    # 若首次读取表格，则导入表格所有内容
                    sheet_part = xls_sheet
                else:
                    # 若非首次读取表格，则将表格后三列数据与之前的表格合并即可
                    sheet_part = xls_sheet.iloc[:, -3:]
                comb_table.append(sheet_part)

                # yield '一共{:d}个表格，已合并{:d}个表格'.format((FIELD_NUMBER + 1) * DOCTYPE_NUMBER, i * DOCTYPE_NUMBER + j + 1)
                yield 'plus'

            field_sum = sheet_part.shape[0]
            sheet_part_plus = pd.DataFrame(
                {'FIELD': pd.Series(cata for _ in range(field_sum)),
                 'FIELD Sum': pd.Series(field_sum for _ in range(field_sum)),
                 '占比': pd.Series((t + 1) / field_sum for t in range(field_sum)),
                 '分档': pd.Series(self.level_judge((t + 1) / field_sum) for t in range(field_sum)),
                 })

            comb_table.append(sheet_part_plus)
            comb_field = pd.concat(comb_table, axis=1)
            comb_field = pd.merge(comb_field, self.contrast_table, how='left', left_on='Institutions', right_on='Institutions')
            comb_all_list.append(comb_field)

        yield "global_finished"
        comb_all = pd.concat(comb_all_list)
        comb_all.rename(columns={'Unnamed: 0': 'Rank'}, inplace=True)

        yield "正在将Global数据写入excel..."
        writer = pd.ExcelWriter(self.resultXlsFile, engine='openpyxl')
        comb_all.to_excel(writer, sheet_name='Global', index=False)
        yield "Global工作表数据已写入excel"
        time.sleep(0.5)

        yield "正在将中国高校数据写入excel..."
        comb_china = comb_all[comb_all['Countries/Regions'] == 'CHINA MAINLAND']
        comb_china = comb_china[comb_china['高校'] != '-']
        comb_china.to_excel(writer, sheet_name='中国高校', index=False)
        yield "中国高校工作表数据已写入excel"

        time.sleep(0.5)

        yield "正在将中国高校统计数据写入excel..."
        self.datas_china = comb_china.reset_index(drop=True)
        comb_china_statistic = self.get_china_statistic()
        comb_china_statistic.to_excel(writer, sheet_name='中国高校统计', index=False)
        yield "中国高校统计表已写入excel"
        writer.close()

        # 将中国高校统计的“上海交通大学”一行标红
        yield "正在将上海交通大学对应数据标红..."
        resESI = op.load_workbook(self.resultXlsFile)
        tableChina = resESI['中国高校统计']
        fontSJTU = op.styles.Font(bold=True, color='FF0000')
        for row in range(10):
            if tableChina.cell(row=row + 1, column=1).value == '上海交通大学':
                for i in range(6):
                    tableChina.cell(row=row + 1, column=i + 1).font = fontSJTU
                break
        resESI.save(self.resultXlsFile)

        yield "数据已全部写入excel！"
        yield "excel文件路径：{:s}".format(self.resultXlsFile)
        yield "finished_com"

    def get_china_statistic(self):
        """
        统计中国高校数据
        """
        df_rank = self.datas_china[self.datas_china['FIELD'] == 'ALL']
        df_rank = df_rank.drop(df_rank[df_rank['高校'] == '-'].index)
        df_rank = df_rank.loc[:, ['高校', 'Rank']]
        df_rank.rename(columns={'Rank': '总名次'}, inplace=True)

        self.datas_china.drop(self.datas_china[self.datas_china['FIELD'] == 'ALL'].index, inplace=True)
        self.datas_china.drop(self.datas_china[self.datas_china['高校'] == '-'].index, inplace=True)

        level_list = ['百分之一', '千分之一', '万分之一', '总计']

        for level in level_list:
            if level == '总计':
                df_level = self.datas_china
            else:
                df_level = self.datas_china[self.datas_china['分档'] == level]
            df_counts = pd.DataFrame(df_level['高校'].value_counts())
            df_counts = df_counts.reset_index()
            df_counts.rename(columns={'高校': level, 'index': '高校'}, inplace=True)
            if level == '百分之一':
                df_result = df_counts
            else:
                df_result = pd.merge(df_result, df_counts, how='outer', on=['高校']).fillna(0)
                df_result = df_result.astype({level: 'int64'})

        df_result = pd.merge(df_result, df_rank, how='left', on=['高校']).fillna(0)
        df_result = df_result.astype({'总名次': 'int64'})
        df_result.sort_values(by=['总名次'], inplace=True)
        return df_result

    def get_excel_filename(self):
        """
        返回合并后的excel文件路径
        """
        return self.resultXlsFile

    def level_judge(self, index):
        """
        判断学科级别
        """
        if index > 0.1:
            return '百分之一'
        elif index > 0.01:
            return '千分之一'
        else:
            return '万分之一'
