# -*- coding: utf-8 -*-
"""
Created on Mon Jan 25 11:38:35 2021

@author: whyzzh
"""
import numpy as np
import pandas as pd
import openpyxl as op
from openpyxl.styles import Border, Side, colors, Alignment


def get_seri(Data_spe, Target_data, loc_type, data_columns):
    Data_spe.columns = [date_str]
    # 将'百分之一'数据按照结果统计表顺序排序
    '''
    Target_data = Target_data_1
    data_columns = data_columns_1
    '''
    ESI_spe = pd.concat([Target_data.iloc[:, 0], Target_data.iloc[:, 1:data_columns]], axis=1)
    ESI_spe_new = pd.merge(left=ESI_spe, right=Data_spe, how='left', on=loc_type)
    ESI_spe_new.iloc[-1, -1] = int(np.sum(ESI_spe_new[date_str]))
    spe_seri = ESI_spe_new.iloc[:, -1]
    return spe_seri


def font_border_modi(ws_total, hund_seri, col_end_hund, col_end_thou, col_end_tenthou, Target_data):
    # 定义字体
    font_data = op.styles.Font(name='Times New Roman', size=11, color='000000')
    font_data_b = op.styles.Font(name='Times New Roman', size=11, bold=True, color='000000')
    font_data_red = op.styles.Font(name='Times New Roman', size=11, color='FF0000')
    font_data_red_b = op.styles.Font(name='Times New Roman', size=11, bold=True, color='FF0000')
    # 定义边框
    border_set = Border(left=Side(style='thin', color=colors.BLACK),
                        right=Side(style='thin', color=colors.BLACK),
                        top=Side(style='thin', color=colors.BLACK),
                        bottom=Side(style='thin', color=colors.BLACK))
    # 定义对齐格式
    alignment_set = Alignment(horizontal='center', vertical='center')

    # 修改格式
    ws_total.cell(row=2, column=col_end_hund + 1).font = font_data_red_b
    ws_total.cell(row=2, column=col_end_thou + 2).font = font_data_red_b
    ws_total.cell(row=2, column=col_end_tenthou + 3).font = font_data_red_b

    ws_total.cell(row=2, column=col_end_hund).font = font_data_b
    ws_total.cell(row=2, column=col_end_thou + 1).font = font_data_b
    ws_total.cell(row=2, column=col_end_tenthou + 2).font = font_data_b

    ws_total.cell(row=2, column=col_end_hund + 1).border = border_set
    ws_total.cell(row=2, column=col_end_thou + 2).border = border_set
    ws_total.cell(row=2, column=col_end_tenthou + 3).border = border_set

    ws_total.cell(row=2, column=col_end_hund + 1).alignment = alignment_set
    ws_total.cell(row=2, column=col_end_thou + 2).alignment = alignment_set
    ws_total.cell(row=2, column=col_end_tenthou + 3).alignment = alignment_set

    for i in range(len(hund_seri)):
        if Target_data.iloc[i, 0] == '上海' or Target_data.iloc[i, 0] == '总计':
            font_red_f = font_data_red_b
            font_f = font_data_b
        else:
            font_red_f = font_data_red
            font_f = font_data

        ws_total.cell(row=i + 3, column=col_end_hund + 1).font = font_red_f
        ws_total.cell(row=i + 3, column=col_end_thou + 2).font = font_red_f
        ws_total.cell(row=i + 3, column=col_end_tenthou + 3).font = font_red_f

        ws_total.cell(row=i + 3, column=col_end_hund).font = font_f
        ws_total.cell(row=i + 3, column=col_end_thou + 1).font = font_f
        ws_total.cell(row=i + 3, column=col_end_tenthou + 2).font = font_f

        ws_total.cell(row=i + 3, column=col_end_hund + 1).border = border_set
        ws_total.cell(row=i + 3, column=col_end_thou + 2).border = border_set
        ws_total.cell(row=i + 3, column=col_end_tenthou + 3).border = border_set


def data_update(Year, Month, filename_target, Target_data, ESI_data, loc_type, data_columns, sheetname):
    '''
    filename_target
    Target_data = Target_data_2 : 通过pandas读取的数据表
    ESI_data = ESIShanghai : 需要进行更新的目标工作表
    loc_type = table_title_2 : 指提取数据表中的“高校”一列做统计
    data_columns = data_columns_2 : 需要新增列的表头
    sheetname = '上海高校'
    '''
    # 统计出各个省份的'百分之一'学科数据，并去掉异常值
    Data_hund = pd.DataFrame(pd.pivot_table(ESI_data, index=loc_type, aggfunc='size'))
    Data_thou = pd.DataFrame(pd.pivot_table(ESI_data[ESI_data['分档'] != '百分之一'], index=loc_type, aggfunc='size'))
    Data_tenthou = pd.DataFrame(pd.pivot_table(ESI_data[ESI_data['分档'] == '万分之一'], index=loc_type, aggfunc='size'))
    hund_seri = get_seri(Data_hund, Target_data, loc_type, data_columns)
    thou_seri = get_seri(Data_thou, Target_data, loc_type, data_columns)
    tenthou_seri = get_seri(Data_tenthou, Target_data, loc_type, data_columns)

    # 打开需要修改的统计结果表

    ws_total = wb_modi.get_sheet_by_name(sheetname)

    # 取消合并第一行单元格
    col_end_hund = data_columns
    col_end_thou = col_end_hund + data_columns - 1
    col_end_tenthou = col_end_thou + data_columns - 1
    ws_total.unmerge_cells(start_row=1, start_column=2, end_row=1, end_column=col_end_hund)
    ws_total.unmerge_cells(start_row=1, start_column=col_end_hund + 1, end_row=1, end_column=col_end_thou)
    ws_total.unmerge_cells(start_row=1, start_column=col_end_thou + 1, end_row=1, end_column=col_end_tenthou)

    # 插入相应的更新列
    ws_total.insert_cols(idx=col_end_tenthou + 1)
    ws_total.insert_cols(idx=col_end_thou + 1)
    ws_total.insert_cols(idx=col_end_hund + 1)

    # 插入相应的表头,并修改字体
    ws_total.cell(row=2, column=col_end_hund + 1).value = date_str
    ws_total.cell(row=2, column=col_end_thou + 2).value = date_str
    ws_total.cell(row=2, column=col_end_tenthou + 3).value = date_str

    # 插入相应的数值
    for i in range(len(hund_seri)):
        ws_total.cell(row=i + 3, column=col_end_hund + 1).value = hund_seri[i]
        ws_total.cell(row=i + 3, column=col_end_thou + 2).value = thou_seri[i]
        ws_total.cell(row=i + 3, column=col_end_tenthou + 3).value = tenthou_seri[i]

    # 字体修改
    font_border_modi(ws_total, hund_seri, col_end_hund, col_end_thou, col_end_tenthou, Target_data)

    # 表头单元格合并
    ws_total.merge_cells(start_row=1, start_column=2, end_row=1, end_column=col_end_hund + 1)
    ws_total.merge_cells(start_row=1, start_column=col_end_hund + 2, end_row=1, end_column=col_end_thou + 2)
    ws_total.merge_cells(start_row=1, start_column=col_end_thou + 3, end_row=1, end_column=col_end_tenthou + 3)

    # 保存文档
    # wb_modi.save(r'D:\Work_Study\ESI_update_test\结果-shanghai.xlsx')
    wb_modi.save(filename_target)


if __name__ == '__main__':
    # 将要添加的年份及月份
    Year = 2021
    Month = 11

    # 原数据表名称路径
    filename_data = r'D:\Work_Study\ESI数据\%d_%02d\ESI-%d%02d.xlsx' % (Year, Month, Year, Month)
    # 统计结果表名称路径
    filename_target = r'D:\Work_Study\ESI数据\%d_%02d\201609-%d%02d ESI统计学科数（发市教委） - 副本.xlsx' % (Year, Month, Year, Month)

    data_columns_1 = (Year - 2017) * 6 + 3 + int((Month + 1) / 2)
    data_columns_2 = (Year - 2017) * 6 + 1 + int((Month + 1) / 2)
    date_str = str(Year) + str(Month).zfill(2)

    # 读取原数据表及统计结果表
    ESIdata = pd.read_excel(filename_data, sheet_name='Global')
    wb_modi = op.load_workbook(filename_target)

    # 读取需要更新的'分省份'统计结果表
    table_title_1 = '省份'
    Target_data_1 = pd.read_excel(filename_target, sheet_name='分省份', skiprows=1, skipfooter=3)
    # 将统计结果表第一列表头改为'省份'
    columns_list_1 = list(Target_data_1.columns)
    columns_list_1[0] = table_title_1
    Target_data_1.columns = columns_list_1
    ESIChina = ESIdata.loc[(ESIdata['Countries/Regions'] == 'CHINA MAINLAND') & (ESIdata['FIELD'] != 'ALL')]
    # 更新'分省份'sheet
    data_update(Year, Month, filename_target, Target_data_1, ESIChina, table_title_1, data_columns_1, '分省份')

    # 读取需要更新的'上海高校'统计结果表
    table_title_2 = '高校'
    Target_data_2 = pd.read_excel(filename_target, sheet_name='上海高校', skiprows=1, skipfooter=3)
    # 将统计结果表第一列表头改为'上海高校'
    columns_list_2 = list(Target_data_2.columns)
    columns_list_2[0] = table_title_2
    Target_data_2.columns = columns_list_2
    ESIShanghai = ESIdata.loc[(ESIdata['省份'] == '上海') & (ESIdata['FIELD'] != 'ALL')]
    # 更新'上海高校'sheet
    data_update(Year, Month, filename_target, Target_data_2, ESIShanghai, table_title_2, data_columns_2, '上海高校')
'''    
def prov_modi(prov):
    
    table_top = int((Year-2020)*6*21+(Month+2-1)/2*21+321)
    table_bottom = table_top+20
        
    wb_modi = op.load_workbook(filename_target)
    ws_shanghai = wb_modi.get_sheet_by_name(prov)
'''
